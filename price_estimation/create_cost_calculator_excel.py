"""
Create an interactive Excel-based cost calculator with all configurations in one sheet.

Usage:
  python create_cost_calculator_excel.py --summary price_estimation/summary_pages.json
"""
import argparse
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_cost_calculator_excel(summary_path: Path, output_path: Path):
    """Create an Excel workbook with single-sheet cost calculator."""
    
    # Load summary data
    with summary_path.open('r', encoding='utf-8') as fh:
        summary = json.load(fh)
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create single consolidated sheet
    create_consolidated_sheet(wb, summary)
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Excel calculator created: {output_path}")
    print("\nHow to use:")
    print("  1. Open the Excel file")
    print("  2. Modify BLUE cells (configurable parameters)")
    print("  3. Results will automatically update")
    print("\nBlue cells sections:")
    print("  - General Parameters (total pages, tokens, exchange rate)")
    print("  - Model Pricing (input/output costs, rate limits)")


def create_consolidated_sheet(wb, summary):
    """Create single sheet with all configuration and results."""
    ws = wb.create_sheet("Cost Calculator", 0)
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    editable_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    editable_font = Font(color="000000", size=10)
    
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    
    result_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    row = 1
    
    # ========== TITLE ==========
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "PDF to Markdown Conversion - Cost Calculator"
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    row += 2
    
    # ========== SECTION 1: GENERAL PARAMETERS ==========
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "GENERAL PARAMETERS (Editable)"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    ws[f'B{row}'].border = thin_border
    row += 1
    
    # Parameters
    params = [
        ('Total Pages', summary['total_pages'], 'pages'),
        ('Tokens per Image (Input)', 258, 'tokens'),
        ('Tokens per Page (Output)', 125, 'tokens'),
        ('Prompt Overhead per Request', 300, 'tokens'),
        ('USD to INR Exchange Rate', 89, 'INR'),
        ('Max Concurrent Requests', 4, 'requests'),
        ('Batch Size', 20, 'files'),
    ]
    
    param_start_row = row
    for i, (label, value, unit) in enumerate(params):
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(size=10)
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = editable_fill
        ws[f'B{row}'].font = editable_font
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = unit
        ws[f'C{row}'].font = Font(size=9, italic=True, color="666666")
        ws[f'C{row}'].border = thin_border
        row += 1
    
    # Computed values
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "COMPUTED VALUES"
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell.border = thin_border
    ws[f'B{row}'].border = thin_border
    row += 1
    
    total_pages_cell = f'B{param_start_row}'
    tokens_per_image_cell = f'B{param_start_row + 1}'
    tokens_per_page_cell = f'B{param_start_row + 2}'
    prompt_overhead_cell = f'B{param_start_row + 3}'
    usd_to_inr_cell = f'B{param_start_row + 4}'
    
    computed_start_row = row
    ws[f'A{row}'] = 'Total Requests (1 per page)'
    ws[f'B{row}'] = f'={total_pages_cell}'
    ws[f'C{row}'] = 'requests'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(size=10)
        ws[f'{col}{row}'].border = thin_border
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    ws[f'C{row}'].font = Font(size=9, italic=True, color="666666")
    row += 1
    
    ws[f'A{row}'] = 'Input Tokens (pages×258 + requests×300)'
    ws[f'B{row}'] = f'={total_pages_cell}*{tokens_per_image_cell}+{total_pages_cell}*{prompt_overhead_cell}'
    ws[f'C{row}'] = 'tokens'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(size=10)
        ws[f'{col}{row}'].border = thin_border
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    ws[f'C{row}'].font = Font(size=9, italic=True, color="666666")
    row += 1
    
    ws[f'A{row}'] = 'Output Tokens (pages×125)'
    ws[f'B{row}'] = f'={total_pages_cell}*{tokens_per_page_cell}'
    ws[f'C{row}'] = 'tokens'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(size=10)
        ws[f'{col}{row}'].border = thin_border
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    ws[f'C{row}'].font = Font(size=9, italic=True, color="666666")
    
    total_requests_cell = f'B{computed_start_row}'
    input_tokens_cell = f'B{computed_start_row + 1}'
    output_tokens_cell = f'B{computed_start_row + 2}'
    
    row += 2
    
    # ========== SECTION 2: MODEL PRICING ==========
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "MODEL PRICING (Editable)"
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # Headers
    headers = ['Model', 'Input (USD/1M)', 'Output (USD/1M)', 'RPM', 'TPM', 'RPD']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    row += 1
    
    # Model data
    models = [
        ('Gemini Free', 0, 0, 15, 1000000, 1500),
        ('Gemini Tier 1', 0.000075, 0.00030, 2000, 4000000, None),
        ('GPT-5', 2.50, 10.00, 500, 2000000, None),
        ('GPT-4.1', 5.00, 15.00, 10000, 30000000, None),
    ]
    
    model_start_row = row
    for model_name, input_cost, output_cost, rpm, tpm, rpd in models:
        ws[f'A{row}'] = model_name
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = input_cost
        ws[f'B{row}'].fill = editable_fill
        ws[f'B{row}'].font = editable_font
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].number_format = '0.000000'
        ws[f'B{row}'].border = thin_border
        
        ws[f'C{row}'] = output_cost
        ws[f'C{row}'].fill = editable_fill
        ws[f'C{row}'].font = editable_font
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        ws[f'C{row}'].number_format = '0.000000'
        ws[f'C{row}'].border = thin_border
        
        ws[f'D{row}'] = rpm
        ws[f'D{row}'].fill = editable_fill
        ws[f'D{row}'].font = editable_font
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'D{row}'].border = thin_border
        
        ws[f'E{row}'] = tpm
        ws[f'E{row}'].fill = editable_fill
        ws[f'E{row}'].font = editable_font
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'E{row}'].border = thin_border
        
        ws[f'F{row}'] = rpd if rpd else ''
        ws[f'F{row}'].fill = editable_fill
        ws[f'F{row}'].font = editable_font
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].border = thin_border
        
        row += 1
    
    row += 2
    
    # ========== SECTION 3: RESULTS ==========
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "COST ESTIMATION RESULTS"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    # Results headers
    result_headers = ['Model', 'Input Cost (USD)', 'Output Cost (USD)', 'Total (USD)', 'Total (INR)', 'Time (min)', 'Time (hrs)']
    for col_idx, header in enumerate(result_headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    row += 1
    
    # Calculate results for each model
    for i, (model_name, _, _, _, _, _) in enumerate(models):
        model_row = model_start_row + i
        
        ws[f'A{row}'] = model_name
        ws[f'A{row}'].font = Font(size=10, bold=True)
        ws[f'A{row}'].fill = result_fill
        ws[f'A{row}'].border = thin_border
        
        # Input Cost (USD) = (input_tokens / 1,000,000) * input_price
        ws[f'B{row}'] = f'={input_tokens_cell}/1000000*B{model_row}'
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].fill = result_fill
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].border = thin_border
        
        # Output Cost (USD) = (output_tokens / 1,000,000) * output_price
        ws[f'C{row}'] = f'={output_tokens_cell}/1000000*C{model_row}'
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].fill = result_fill
        ws[f'C{row}'].alignment = Alignment(horizontal='right')
        ws[f'C{row}'].border = thin_border
        
        # Total (USD)
        ws[f'D{row}'] = f'=B{row}+C{row}'
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].fill = result_fill
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].border = thin_border
        
        # Total (INR)
        ws[f'E{row}'] = f'=D{row}*{usd_to_inr_cell}'
        ws[f'E{row}'].number_format = '₹#,##0.00'
        ws[f'E{row}'].fill = result_fill
        ws[f'E{row}'].font = Font(bold=True)
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].border = thin_border
        
        # Time (minutes) = MAX(input_tokens+output_tokens / TPM, total_requests / RPM)
        ws[f'F{row}'] = f'=MAX(({input_tokens_cell}+{output_tokens_cell})/E{model_row},{total_requests_cell}/D{model_row})'
        ws[f'F{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].fill = result_fill
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].border = thin_border
        
        # Time (hours)
        ws[f'G{row}'] = f'=F{row}/60'
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'G{row}'].fill = result_fill
        ws[f'G{row}'].alignment = Alignment(horizontal='right')
        ws[f'G{row}'].border = thin_border
        
        row += 1
    
    row += 2
    
    # ========== SECTION 4: FORMULAS REFERENCE ==========
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = "FORMULAS REFERENCE"
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.border = thin_border
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = thin_border
    row += 1
    
    formulas = [
        ('Total Requests', '= Total Pages (1 request per page)'),
        ('Input Tokens', '= Pages × Tokens_per_Image + Requests × Prompt_Overhead'),
        ('Output Tokens', '= Pages × Tokens_per_Page'),
        ('Input Cost (USD)', '= (Input Tokens / 1,000,000) × Input Price (USD/1M)'),
        ('Output Cost (USD)', '= (Output Tokens / 1,000,000) × Output Price (USD/1M)'),
        ('Total Cost (USD)', '= Input Cost + Output Cost'),
        ('Total Cost (INR)', '= Total Cost (USD) × USD_to_INR'),
        ('Time (minutes)', '= MAX(Total_Tokens / TPM, Total_Requests / RPM)'),
        ('Time (hours)', '= Time (minutes) / 60'),
    ]
    
    for label, formula in formulas:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(size=9, bold=True)
        ws[f'A{row}'].border = thin_border
        
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = formula
        ws[f'B{row}'].font = Font(size=9, color="333333")
        ws[f'B{row}'].border = thin_border
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].border = thin_border
        row += 1
    
    # Freeze panes at top
    ws.freeze_panes = 'A4'


def main():
    parser = argparse.ArgumentParser(
        description='Create interactive Excel cost calculator for PDF to Markdown conversion'
    )
    parser.add_argument(
        '--summary',
        type=Path,
        default=Path('price_estimation/summary_pages.json'),
        help='Path to summary JSON file (default: price_estimation/summary_pages.json)'
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=Path('price_estimation/cost_calculator.xlsx'),
        help='Output Excel file path (default: price_estimation/cost_calculator.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Validate input file
    if not args.summary.exists():
        print(f"✗ Error: Summary file not found: {args.summary}")
        print(f"  Please run calculate_costs.py first to generate the summary file.")
        return 1
    
    # Create output directory if needed
    args.output.parent.mkdir(parents=True, exist_ok=True)
    
    # Create calculator
    create_cost_calculator_excel(args.summary, args.output)
    
    return 0


if __name__ == '__main__':
    exit(main())
